import requests
import json
import openpyxl as op
import time


requests.DEFAULT_RETRIES=5
s= requests.session()
s.keep_alive=False


ws=op.Workbook()
wb = ws.create_sheet(index=0)
wb.cell(row=1,column=1,value='商品名称')
wb.cell(row=1,column=2,value='品牌')
wb.cell(row=1,column=3,value='价格')
headers = {
	'Cookie': 'vip_cps_cuid=CU1676115410598550179aa54450a6c0; vip_cps_cid=1676115410600_196e5f9f827de1d59b2c0e02b232029f; cps_share=cps_share; PAPVisitorId=472c653004562ca65b962a8e563f2dac; vip_new_old_user=1; vip_city_name=%E5%B9%BF%E5%B7%9E%E5%B8%82; user_class=a; mars_cid=1676115414402_a7342faa478349a4d9384c3962f40178; mars_pid=0; VipUINFO=luc%3Aa%7Csuc%3Aa%7Cbct%3Ac_new%7Chct%3Ac_new%7Cbdts%3A0%7Cbcts%3A0%7Ckfts%3A0%7Cc10%3A0%7Crcabt%3A0%7Cp2%3A0%7Cp3%3A1%7Cp4%3A0%7Cp5%3A0%7Cul%3A3105; vip_address=%257B%2522pname%2522%253A%2522%255Cu6c5f%255Cu82cf%255Cu7701%2522%252C%2522pid%2522%253A%2522103102%2522%252C%2522cname%2522%253A%2522%255Cu5e7f%255Cu5dde%255Cu5e02%2522%252C%2522cid%2522%253A%2522103102101%2522%257D; vip_province=103102; vip_province_name=%E6%B1%9F%E8%8B%8F%E7%9C%81; vip_city_code=103102101; vip_wh=VIP_SH; vip_ipver=31; cps=adp%3Antq8exyc%3A%40_%401676116467372%3Amig_code%3A4f6b50bf15bfa39639d85f5f1e15b10f%3Aac014miuvl0000b5sq8cc2oyplrsivsi; mars_sid=3cf42e14882e4babff997b54232d497e; visit_id=4C8383DDD7815413DA9DB4EA5674417F; pg_session_no=42; vip_tracker_source_from=%7B%22activity_data%22%3A%22%257B%2522biz_data%2522%253A%257B%2522target_type%2522%253A%2522autolist%2522%252C%2522sequence%2522%253A1%252C%2522target_id%2522%253A%252258811586%2522%257D%252C%2522ext_data%2522%253A%257B%2522first_category_id%2522%253A%252230074%2522%252C%25222nd_cat_title%2522%253A%2522%25E5%25A5%25B3%25E5%25A3%25AB%25E8%25A3%25A4%25E8%25A3%2585%2522%252C%2522second_category_id%2522%253A%252252020713%2522%252C%25223rd_cat_title%2522%253A%2522%25E8%25BF%259E%25E4%25BD%2593%25E8%25A3%25A4%2522%252C%25221st_cat_title%2522%253A%2522%25E5%25A5%25B3%25E8%25A3%2585%252F%25E7%2594%25B7%25E8%25A3%2585%252F%25E5%2586%2585%25E8%25A1%25A3%2522%252C%2522is_freestyle%2522%253A%25220%2522%257D%252C%2522obj_data%2522%253A%257B%2522source%2522%253A%2522lighart%2522%252C%2522id%2522%253A%2522GlobalClassify_Category%2522%257D%252C%2522click_fu%2522%253A1%257D%22%2C%22activity_id%22%3A%22lightart_slider%22%2C%22activity_type%22%3A%22interface%22%7D; vip_access_times=%7B%22list%22%3A3%7D',
	'Referer': 'https://category.vip.com/suggest.php?keyword=%E6%8A%A4%E8%82%A4&ff=235|12|1|1',
	'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
	'Connection': 'close'
}


dress=[[['纯色美裙'], ['54886020']], [['高腰美裙'], ['54886032']], [['半截裙'], ['55521161']], [['中长款裙'], ['54886023']], [['碎花美裙'], ['54886027']], [['百褶美裙'], ['54886022']], [['格纹美裙'], ['54886016']], [['吊带美裙'], ['54886031']], [['真丝美裙'], ['54886021']], [['旗袍美裙'], ['54886017']], [['牛仔美裙'], ['54886028']], [['大码裙装'], ['54879403']]]
hot=[[['加厚保暖'], ['62221605']], [['冬季尚新'], ['61971038']], [['针织衫'], ['53986298']], [['唯品独家'], ['61971053']], [['休闲裤'], ['53986312']], [['连衣裙'], ['53986307']], [['女式毛衣'], ['55515022']], [['外套'], ['53986308']], [['牛仔裤'], ['53986316']], [['衬衫'], ['53986299']], [['T恤'], ['53986306']], [['套装'], ['53986294']], [['半截裙'], ['53986314']], [['妈妈装'], ['53986804']]]
coat=[[['衬衫'], ['53986299']], [['T恤'], ['53986306']], [['外套'], ['53986308']], [['针织衫'], ['53986298']], [['风衣'], ['53986296']], [['西服'], ['53986315']], [['卫衣'], ['53986297']], [['马夹'], ['53986639']], [['打底衫'], ['59968549']], [['背心'], ['53986303']], [['大衣'], ['53986305']], [['皮衣和皮草'], ['53986295']], [['毛衣'], ['55515022']], [['羽绒服'], ['53986313']]]
trousers=[[['运动裤'], ['58811183']], [['西裤'], ['55518075']], [['连体裤'], ['58811586']], [['打底裤'], ['53986311']], [['休闲裤'], ['53986312']], [['牛仔裤'], ['53986316']], [['短裤'], ['53986751']], [['直筒裤'], ['55516985']], [['工装裤'], ['55518074']], [['小脚裤'], ['55517801']], [['阔腿裤'], ['55516984']], [['哈伦裤'], ['55517076']], [['半截裙'], ['53986314']]]
underwear=[[['纯棉内裤'], ['59495926']], [['女士家居服'], ['59495911']], [['纯棉家居服'], ['59495921']], [['塑身内衣'], ['59495929']], [['女士睡袍'], ['58325065']], [['美背文胸'], ['59495679']], [['薄杯文胸'], ['59493551']], [['运动文胸'], ['59495397']], [['女士内裤'], ['59450382']], [['无钢圈'], ['59450613']], [['聚拢'], ['59450718']], [['女袜'], ['59450379']], [['裤袜'], ['59450380']], [['女士吊带/打底背心'], ['59450384']], [['文胸套装'], ['59450750']]]

lady=[dress,hot,coat,trousers,underwear]

# url = 'https://mapi.vip.com/vips-mobile/rest/shopping/pc/product/list/rank/rule/v2?callback=getProductIdsListRank&app_name=shop_pc&app_version=4.0&warehouse=VIP_SH&fdc_area_id=103102101&client=pc&mobile_platform=1&province_id=103102&api_key=70f71280d5d547b2a7bb370a529aeea1&user_id=&mars_cid=1676115414402_a7342faa478349a4d9384c3962f40178&wap_consumer=a&uid=&abtestId=1872&mtmsRuleId=54886032&scene=rule_stream&sizeNames=&props=&vipService=&bigSaleTagIds=&filterStock=0&brandStoreSns=&sort=0&pageOffset=0&salePlatform=1&_=1676118708610'
# html=requests.get(url,headers=headers)
# print(html.text)

for y in range(2,4):
	for z in range(3,10):
		n = 50
		num = 2
		for x in range(0,(n+1)*120,120):


			t=time.time()
			T=int(round(t*1000))-50
			url=f'https://mapi.vip.com/vips-mobile/rest/shopping/pc/product/list/rank/rule/v2?callback=getProductIdsListRank&app_name=shop_pc&app_version=4.0&warehouse=VIP_SH&fdc_area_id=103102101&client=pc&mobile_platform=1&province_id=103102&api_key=70f71280d5d547b2a7bb370a529aeea1&user_id=&mars_cid=1676115414402_a7342faa478349a4d9384c3962f40178&wap_consumer=a&uid=&abtestId=1872&mtmsRuleId={lady[y][z][1][0]}&scene=rule_stream&sizeNames=&props=&vipService=&bigSaleTagIds=&filterStock=0&brandStoreSns=&sort=0&pageOffset={num}&salePlatform=1&_={T}'
			#纯色美裙https://mapi.vip.com/vips-mobile/rest/shopping/pc/product/list/rank/rule/v2?callback=getProductIdsListRank&app_name=shop_pc&app_version=4.0&warehouse=VIP_SH&fdc_area_id=103102101&client=pc&mobile_platform=1&province_id=103102&api_key=70f71280d5d547b2a7bb370a529aeea1&user_id=&mars_cid=1676115414402_a7342faa478349a4d9384c3962f40178&wap_consumer=a&uid=&abtestId=1872&mtmsRuleId=54886020&scene=rule_stream&sizeNames=&props=&vipService=&bigSaleTagIds=&filterStock=0&brandStoreSns=&sort=0&pageOffset=0&salePlatform=1&_=1676125104884
			#中长款裙https://mapi.vip.com/vips-mobile/rest/shopping/pc/product/list/rank/rule/v2?callback=getProductIdsListRank&app_name=shop_pc&app_version=4.0&warehouse=VIP_SH&fdc_area_id=103102101&client=pc&mobile_platform=1&province_id=103102&api_key=70f71280d5d547b2a7bb370a529aeea1&user_id=&mars_cid=1676115414402_a7342faa478349a4d9384c3962f40178&wap_consumer=a&uid=&abtestId=1872&mtmsRuleId=54886023&scene=rule_stream&sizeNames=&props=&vipService=&bigSaleTagIds=&filterStock=0&brandStoreSns=&sort=0&pageOffset=0&salePlatform=1&_=1676125291042
			#半截裙https://mapi.vip.com/vips-mobile/rest/shopping/pc/product/list/rank/rule/v2?callback=getProductIdsListRank&app_name=shop_pc&app_version=4.0&warehouse=VIP_SH&fdc_area_id=103102101&client=pc&mobile_platform=1&province_id=103102&api_key=70f71280d5d547b2a7bb370a529aeea1&user_id=&mars_cid=1676115414402_a7342faa478349a4d9384c3962f40178&wap_consumer=a&uid=&abtestId=1872&mtmsRuleId=55521161&scene=rule_stream&sizeNames=&props=&vipService=&bigSaleTagIds=&filterStock=0&brandStoreSns=&sort=0&pageOffset=0&salePlatform=1&_=1676126877608
			#13碎花美裙https://mapi.vip.com/vips-mobile/rest/shopping/pc/product/list/rank/rule/v2?callback=getProductIdsListRank&app_name=shop_pc&app_version=4.0&warehouse=VIP_SH&fdc_area_id=103102101&client=pc&mobile_platform=1&province_id=103102&api_key=70f71280d5d547b2a7bb370a529aeea1&user_id=&mars_cid=1676115414402_a7342faa478349a4d9384c3962f40178&wap_consumer=a&uid=&abtestId=1872&mtmsRuleId=54886027&scene=rule_stream&sizeNames=&props=&vipService=&bigSaleTagIds=&filterStock=0&brandStoreSns=&sort=0&pageOffset=0&salePlatform=1&_=1676127111194





			html=requests.get(url, headers=headers)
			#print(html.text)

			start=html.text.index('{')
			end=html.text.index('})')+1
			json_data=json.loads(html.text[start:end])
			#print(json_data)
			#print(json_data['data']['products'])
			#print('\n')
			#print(len(json_data['data']['products']))
			product_ids=json_data['data']['products']


			for product_id in product_ids:
				#print('商品id:',product_id['pid'])
				product_url='https://mapi.vip.com/vips-mobile/rest/shopping/pc/product/module/list/v2?callback=getMerchandiseDroplets1&app_name=shop_pc&app_version=4.0&warehouse=VIP_SH&fdc_area_id=103102101&client=pc&mobile_platform=1&province_id=103102&api_key=70f71280d5d547b2a7bb370a529aeea1&user_id=&mars_cid=1676115414402_a7342faa478349a4d9384c3962f40178&wap_consumer=a&productIds={}%2C&scene=rule_stream&standby_id=nature&extParams=%7B%22stdSizeVids%22%3A%22%22%2C%22preheatTipsVer%22%3A%223%22%2C%22couponVer%22%3A%22v2%22%2C%22exclusivePrice%22%3A%221%22%2C%22iconSpec%22%3A%222x%22%2C%22ic2label%22%3A1%2C%22superHot%22%3A1%2C%22bigBrand%22%3A%221%22%7D&context=&_={}'.format(product_id['pid'],T)
				#https://mapi.vip.com/vips-mobile/rest/shopping/pc/product/module/list/v2?callback=getMerchandiseDroplets1&app_name=shop_pc&app_version=4.0&warehouse=VIP_SH&fdc_area_id=103102101&client=pc&mobile_platform=1&province_id=103102&api_key=70f71280d5d547b2a7bb370a529aeea1&user_id=&mars_cid=1676115414402_a7342faa478349a4d9384c3962f40178&wap_consumer=a&productIds={}%2C&scene=rule_stream&standby_id=nature&extParams=%7B%22stdSizeVids%22%3A%22%22%2C%22preheatTipsVer%22%3A%223%22%2C%22couponVer%22%3A%22v2%22%2C%22exclusivePrice%22%3A%221%22%2C%22iconSpec%22%3A%222x%22%2C%22ic2label%22%3A1%2C%22superHot%22%3A1%2C%22bigBrand%22%3A%221%22%7D&context=&_=1676125104885
				#https://mapi.vip.com/vips-mobile/rest/shopping/pc/product/module/list/v2?callback=getMerchandiseDroplets1&app_name=shop_pc&app_version=4.0&warehouse=VIP_SH&fdc_area_id=103102101&client=pc&mobile_platform=1&province_id=103102&api_key=70f71280d5d547b2a7bb370a529aeea1&user_id=&mars_cid=1676115414402_a7342faa478349a4d9384c3962f40178&wap_consumer=a&productIds={}%2C&scene=rule_stream&standby_id=nature&extParams=%7B%22stdSizeVids%22%3A%22%22%2C%22preheatTipsVer%22%3A%223%22%2C%22couponVer%22%3A%22v2%22%2C%22exclusivePrice%22%3A%221%22%2C%22iconSpec%22%3A%222x%22%2C%22ic2label%22%3A1%2C%22superHot%22%3A1%2C%22bigBrand%22%3A%221%22%7D&context=&_=1676125291043
				#https://mapi.vip.com/vips-mobile/rest/shopping/pc/product/module/list/v2?callback=getMerchandiseDroplets1&app_name=shop_pc&app_version=4.0&warehouse=VIP_SH&fdc_area_id=103102101&client=pc&mobile_platform=1&province_id=103102&api_key=70f71280d5d547b2a7bb370a529aeea1&user_id=&mars_cid=1676117136412_7a36fb645c205ffb3ff52054463ebe57&wap_consumer=a&productIds={}%2C&scene=rule_stream&standby_id=nature&extParams=%7B%22stdSizeVids%22%3A%22%22%2C%22preheatTipsVer%22%3A%223%22%2C%22couponVer%22%3A%22v2%22%2C%22exclusivePrice%22%3A%221%22%2C%22iconSpec%22%3A%222x%22%2C%22ic2label%22%3A1%2C%22superHot%22%3A1%2C%22bigBrand%22%3A%221%22%7D&context=&_=1676125218748


				product_html = requests.get(product_url, headers=headers)
				product_start=product_html.text.index('{')
				product_end=product_html.text.index('})')+1
				product_json_data=json.loads(product_html.text[product_start:product_end])
				product_info_data=product_json_data['data']['products'][0]
				product_title=product_info_data['title']
				product_brand=product_info_data['brandShowName']
				product_price = product_info_data['price']['salePrice']
				#print('商品名称：{}，品牌：{}，折后价格：{}'.format(product_title, product_brand, product_price))
				wb.cell(row=num,column=1,value=product_title)
				wb.cell(row=num, column=2, value=product_brand)
				wb.cell(row=num, column=3, value=product_price)
				num+=1
				#print(product_html.text)
		print(num)
		ws.save('{}.xlsx'.format(lady[y][z][0][0]))
